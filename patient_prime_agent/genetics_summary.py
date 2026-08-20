from __future__ import annotations

import argparse
import json
from collections import OrderedDict
from datetime import datetime
from pathlib import Path
from typing import Any

from .utils import ensure_dir

DEFAULT_INPUT_PATH = Path("patient_data") / "Genetics" / "genetics_data.xlsx"
DEFAULT_OUTPUT_PATH = Path("reports") / "genetics" / "genetics_clinical_summary.json"

REPORT_TYPE = "Pharmacogenomic (PGx) Clinical Interpretation Report"

# The 9 pharmacogenes this report profiles in detail. Chosen because they are the genes in
# this dataset that carry star-allele/diplotype-level annotation (see metabolizer_profile
# workflow in SKILL.md) rather than a plain rs-genotype.
METABOLIZER_GENES = ["CYP2D6", "CYP2C19", "CYP2C9", "CYP2B6", "CYP3A5", "CYP1A2", "CYP3A4", "UGT1A4", "POR"]

# Fixed, documented re-wording of the sheet's own "PGx response" category into readable
# English. This translates the source's own value; it never adds a new clinical claim.
RESPONSE_PHRASES = {
    "efficacy": "favorable efficacy signal",
    "reduced efficacy": "reduced efficacy signal",
    "toxicity": "increased adverse-reaction risk",
    "moderate": "moderate response",
}

# Sheet "Drug Class" -> report therapeutic-class bucket, matching the reference shape.
THERAPEUTIC_CLASS_MAP = {
    "antidepressant": "antidepressants",
    "atypical antipsychotic": "antipsychotics",
    "typical antipsychotic": "antipsychotics",
    "antiepileptic": "mood_stabilizers_antiepileptics",
    "antiepileptics": "mood_stabilizers_antiepileptics",
    "anticonvulsants": "mood_stabilizers_antiepileptics",
    "mood stabilizer": "mood_stabilizers_antiepileptics",
    "stimulant": "adhd_stimulants",
    "non-stimulant adhd medication": "adhd_stimulants",
}
OTHER_CLASS_KEY = "other_panel_entries"
THERAPEUTIC_CLASS_ORDER = [
    "antidepressants",
    "antipsychotics",
    "mood_stabilizers_antiepileptics",
    "adhd_stimulants",
    OTHER_CLASS_KEY,
]

PURPOSE_TEMPLATE = (
    "This report summarizes a pharmacogenomic (PGx) panel that cross-references the patient's "
    "genotype at {variant_count} pharmacologically relevant loci against the gene-drug annotations "
    "present in the source dataset, covering {drug_count} medications. It predicts likely efficacy "
    "and toxicity/side-effect risk; this is a decision-support tool, not a diagnosis. It does not "
    "indicate which condition(s) the patient has."
)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Generate a pharmacogenomic clinical interpretation summary from a genetics data spreadsheet")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT_PATH)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_PATH)
    args = parser.parse_args(argv)

    rows = load_rows(args.input)
    report = build_report(rows)
    ensure_dir(args.output.parent)
    args.output.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote genetics clinical summary to {args.output}")
    return 0


# ----------------------------------------------------------------------
# loading
# ----------------------------------------------------------------------
def load_rows(input_path: Path) -> list[dict[str, Any]]:
    if not input_path.exists():
        return []
    import openpyxl

    workbook = openpyxl.load_workbook(input_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]

    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        record = {headers[index]: values[index] for index in range(len(headers)) if index < len(values)}
        gene = _clean(record.get("Gene"))
        drug = _clean(record.get("Drug"))
        if not gene and not drug:
            continue
        rows.append(
            {
                "gene": gene,
                "gene_upper": gene.upper() if gene else None,
                "drug": drug,
                "drug_title": drug.title() if drug else None,
                "drug_class_raw": _clean(record.get("Drug Class")),
                "disease": _clean(record.get("Disease")),
                "medical_conditions": _clean(record.get("Medical Conditions")),
                "patient_genotype": _clean(record.get("Patient_genotype")),
                "diplotype": _clean(record.get("Diplotype")),
                "phenotype": _clean(record.get("Phenotype")),
                "pgx_response": _clean(record.get("PGx response")),
                "significance": _clean(record.get("Significance")),
                "recommendation": _clean(record.get("Recommendation Statement")),
                "therapeutic_interaction": _clean(record.get("Therapeutic Interaction")),
                "fda_flag": _clean(record.get("FDA")),
                "fda_boxed_warning": _clean(record.get("FDA Boxed Warning")),
            }
        )
    return rows


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _is_significant(row: dict[str, Any]) -> bool:
    return (row.get("significance") or "").strip().lower() == "yes"


# ----------------------------------------------------------------------
# report assembly
# ----------------------------------------------------------------------
def build_report(rows: list[dict[str, Any]]) -> dict[str, Any]:
    if not rows:
        return _empty_report()

    genes = {row["gene_upper"] for row in rows if row["gene_upper"]}
    drugs = {row["drug"] for row in rows if row["drug"]}

    metabolizer_profile = _build_metabolizer_profile(rows)
    findings_by_class = _build_findings_by_class(rows)
    safety_flags = _build_priority_safety_flags(rows)
    boxed_warning_drugs = sorted({row["drug_title"] for row in rows if row["fda_boxed_warning"] and row["drug_title"]})

    return {
        "report_type": REPORT_TYPE,
        "patient": {
            "patient_id": None,
            "report_date": datetime.now().strftime("%Y-%m-%d"),
            "specimen": None,
            "ordering_context": None,
            "variants_analyzed": len(genes),
            "drugs_covered": len(drugs),
        },
        "purpose": PURPOSE_TEMPLATE.format(variant_count=len(genes), drug_count=len(drugs)),
        "metabolizer_profile": metabolizer_profile,
        "findings_by_therapeutic_class": findings_by_class,
        "priority_safety_flags": safety_flags,
        "boxed_warning_note": (
            f"{len(boxed_warning_drugs)} medication(s) in this panel carry an FDA boxed warning in the "
            f"source data ({', '.join(boxed_warning_drugs[:10])}{'...' if len(boxed_warning_drugs) > 10 else ''}). "
            "Standard FDA boxed warnings apply independent of these genetic findings; see each drug's own "
            "prescribing information."
            if boxed_warning_drugs
            else "No FDA boxed warning text was present in the source data for the medications in this panel."
        ),
        "clinical_conclusion": _build_clinical_conclusion(metabolizer_profile, safety_flags),
    }


def _build_metabolizer_profile(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_gene: dict[str, list[dict[str, Any]]] = {gene: [] for gene in METABOLIZER_GENES}
    for row in rows:
        if row["gene_upper"] in by_gene:
            by_gene[row["gene_upper"]].append(row)

    profile: list[dict[str, Any]] = []
    for gene in METABOLIZER_GENES:
        gene_rows = [row for row in by_gene[gene] if row["drug"]]
        if not gene_rows:
            continue

        diplotypes = {row["diplotype"] for row in gene_rows if row["diplotype"]}
        phenotypes = {row["phenotype"] for row in gene_rows if row["phenotype"]}
        genotypes = {row["patient_genotype"] for row in gene_rows if row["patient_genotype"]}

        if diplotypes and phenotypes:
            status = f"{next(iter(diplotypes))} - {next(iter(phenotypes))}"
        elif genotypes:
            status = next(iter(genotypes))
        else:
            continue

        # group this gene's drugs by their PGx-response phrase so the impact sentence reads
        # "DrugA, DrugB - phrase" the same way for every drug that shares that response.
        by_phrase: "OrderedDict[str, list[str]]" = OrderedDict()
        for row in gene_rows:
            phrase = RESPONSE_PHRASES.get((row["pgx_response"] or "").lower())
            if not phrase or not row["drug_title"]:
                continue
            by_phrase.setdefault(phrase, [])
            if row["drug_title"] not in by_phrase[phrase]:
                by_phrase[phrase].append(row["drug_title"])
        impact = "; ".join(f"{', '.join(names)} - {phrase}" for phrase, names in by_phrase.items())

        profile.append({"gene": gene, "status": status, "impact": impact or None})
    return profile


def _build_findings_by_class(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    findings: dict[str, list[dict[str, Any]]] = {key: [] for key in THERAPEUTIC_CLASS_ORDER}
    for row in rows:
        if not row["drug"] or not row["gene"]:
            continue
        class_key = THERAPEUTIC_CLASS_MAP.get((row["drug_class_raw"] or "").lower(), OTHER_CLASS_KEY)
        genetic_basis = row["gene_upper"] or row["gene"]
        if row["patient_genotype"]:
            genetic_basis = f"{genetic_basis} ({row['patient_genotype']})"
        elif row["diplotype"]:
            genetic_basis = f"{genetic_basis} ({row['diplotype']})"
        findings[class_key].append(
            {
                "drug": row["drug_title"],
                "genetic_basis": genetic_basis,
                "predicted_effect": row["pgx_response"],
                "significant": _is_significant(row),
            }
        )
    return {key: value for key, value in findings.items() if value}


def _build_priority_safety_flags(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Software-detected candidates, not an exhaustive clinical review -- see SKILL.md."""

    flags: "OrderedDict[str, dict[str, Any]]" = OrderedDict()
    for row in rows:
        if not row["gene_upper"] or not _is_significant(row):
            continue
        response = (row["pgx_response"] or "").lower()
        if response not in {"toxicity", "reduced efficacy"}:
            continue
        # Only the gene-specific fields are searched here. FDA Boxed Warning text is
        # drug-level labeling (e.g. the standard antidepressant suicidality class warning)
        # repeated across many unrelated rows for the same drug -- including it would flag
        # genes that have nothing to do with the actual finding.
        combined_text = " ".join(filter(None, [row["recommendation"], row["therapeutic_interaction"]])).lower()
        if "suicid" in combined_text:
            severity = "high"
        elif row["gene_upper"] in METABOLIZER_GENES and response == "toxicity":
            severity = "moderate"
        else:
            continue

        key = row["gene_upper"]
        entry = flags.get(key)
        drug = row["drug_title"] or "unspecified medication"
        risk_phrase = RESPONSE_PHRASES.get(response, response)
        if entry is None:
            flags[key] = {
                "flag": f"{row['gene_upper']}" + (f" ({row['patient_genotype']})" if row["patient_genotype"] else ""),
                "risk": f"{risk_phrase.capitalize()} on {drug}",
                "severity": severity,
                "_drugs": [drug],
            }
        elif drug not in entry["_drugs"]:
            entry["_drugs"].append(drug)
            entry["risk"] = f"{risk_phrase.capitalize()} on {', '.join(entry['_drugs'])}"
            if severity == "high":
                entry["severity"] = "high"

    ordered = sorted(flags.values(), key=lambda item: (item["severity"] != "high", item["flag"]))
    for item in ordered:
        item.pop("_drugs", None)
    return ordered[:10]


def _build_clinical_conclusion(metabolizer_profile: list[dict[str, Any]], safety_flags: list[dict[str, Any]]) -> dict[str, Any]:
    reduced_function_genes = [
        entry["gene"]
        for entry in metabolizer_profile
        if any(term in entry["status"].lower() for term in ("intermediate", "poor", "decreased"))
    ]
    high_severity = [flag["flag"] for flag in safety_flags if flag["severity"] == "high"]

    impression_parts = []
    if reduced_function_genes:
        impression_parts.append(
            f"The profile shows reduced-function metabolizer status across {len(reduced_function_genes)} gene(s) "
            f"({', '.join(reduced_function_genes)}), predicting atypical drug clearance for the medications linked "
            "to those genes."
        )
    if high_severity:
        impression_parts.append(
            f"{len(high_severity)} genetic marker(s) ({', '.join(high_severity)}) flag an elevated adverse-effect "
            "risk that should prompt caution or enhanced monitoring for the associated medications."
        )
    if not impression_parts:
        impression_parts.append(
            "No reduced-function metabolizer status or high-severity safety flags were identified from the "
            "extracted panel data."
        )

    return {
        "impression": " ".join(impression_parts),
        "recommendations": [
            {
                "action": "Avoid or use with heightened monitoring",
                "detail": f"See priority_safety_flags for the specific gene-drug pairs ({', '.join(high_severity)})." if high_severity else "No high-severity flags identified.",
                "priority": "high",
            },
            {
                "action": "Dose with caution / consider therapeutic drug monitoring",
                "detail": f"See metabolizer_profile for reduced-function genes ({', '.join(reduced_function_genes)})." if reduced_function_genes else "No reduced-function metabolizer genes identified.",
                "priority": "moderate",
            },
            {
                "action": "Contextualize",
                "detail": "Interpret alongside actual clinical history, current medication list, and labs -- this panel alone does not establish a diagnosis.",
                "priority": "informational",
            },
        ],
    }


def _empty_report() -> dict[str, Any]:
    return {
        "report_type": REPORT_TYPE,
        "patient": {
            "patient_id": None,
            "report_date": datetime.now().strftime("%Y-%m-%d"),
            "specimen": None,
            "ordering_context": None,
            "variants_analyzed": 0,
            "drugs_covered": 0,
        },
        "purpose": PURPOSE_TEMPLATE.format(variant_count=0, drug_count=0),
        "metabolizer_profile": [],
        "findings_by_therapeutic_class": {},
        "priority_safety_flags": [],
        "boxed_warning_note": "No FDA boxed warning text was present in the source data for the medications in this panel.",
        "clinical_conclusion": {
            "impression": "No genetics source data was available to process.",
            "recommendations": [],
        },
    }


if __name__ == "__main__":
    raise SystemExit(main())
